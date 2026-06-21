"""Unattended batch processing of many datasets.

Reads a *manifest* (CSV / TSV / XLSX) listing datasets -- each a base directory
plus an optional config file -- and runs each one through the gliding pipeline.
Designed to run for a long time without supervision:

* a **pre-flight check** validates the whole list before any heavy work starts;
* every dataset runs inside a hardened ``try/except`` (including ``SystemExit``),
  so a failure is logged and the run *continues* to the next dataset;
* a JSON **state file** records each dataset's outcome plus an input/config
  signature, so re-runs **skip** datasets already completed successfully (unless
  their inputs or config changed) -- and the state is saved after every dataset,
  so an interrupted run resumes cleanly;
* **detailed, timestamped logs** are written for the whole run and per dataset.

The gliding pipeline itself is imported lazily, so manifest parsing, the
pre-flight check, and the state logic stay importable without the image stack.
"""
from __future__ import annotations

import csv
import hashlib
import json
import logging
import os
import sys
import time
import traceback
from contextlib import redirect_stderr, redirect_stdout
from dataclasses import dataclass, field
from datetime import datetime
from typing import Dict, List, Optional

LOGGER = logging.getLogger("fastrack.batch")

# Accepted column names (case-insensitive) for each manifest field.
_ALIASES = {
    "base_dir": ["base_dir", "basedir", "directory", "dir", "dataset", "data", "path"],
    "config": ["config", "config_file", "cfg", "toml", "settings"],
    "name": ["name", "id", "label"],
}


@dataclass
class DatasetSpec:
    name: str
    base_dir: str
    config: Optional[str] = None
    #: problems found by pre-flight (empty == ok)
    problems: List[str] = field(default_factory=list)


# --------------------------------------------------------------------------- #
# Manifest parsing
# --------------------------------------------------------------------------- #
def _norm_header(cols):
    """Map actual column names to canonical field names via the alias table."""
    lookup = {}
    for i, col in enumerate(cols):
        key = str(col or "").strip().lower()
        for canon, names in _ALIASES.items():
            if key in names:
                lookup[canon] = i
    return lookup


def _rows_from_csv(path):
    with open(path, newline="") as f:
        sample = f.read(4096)
        f.seek(0)
        if path.lower().endswith(".tsv"):
            dialect = csv.excel_tab
        else:
            try:
                dialect = csv.Sniffer().sniff(sample, delimiters=",\t;")
            except csv.Error:
                dialect = csv.excel
        for row in csv.reader(f, dialect):
            yield row


def _rows_from_xlsx(path):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - only without the extra
        raise ImportError(
            "Reading .xlsx manifests needs openpyxl. Install it with:\n"
            "    pip install 'fastrack[batch]'\n"
            "or export your sheet to CSV and pass that instead."
        ) from exc
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    for row in ws.iter_rows(values_only=True):
        yield list(row)


def read_manifest(path: str) -> List[DatasetSpec]:
    """Parse a manifest into :class:`DatasetSpec` rows.

    Required column: a base directory (``base_dir``/``directory``/``dataset``/...).
    Optional columns: ``config`` (path to a TOML config) and ``name``.  Relative
    paths are resolved against the manifest's own directory so lists are portable.
    Blank rows and rows whose first cell starts with ``#`` are ignored.
    """
    manifest_dir = os.path.dirname(os.path.abspath(path))
    ext = os.path.splitext(path)[1].lower()
    rows = _rows_from_xlsx(path) if ext == ".xlsx" else _rows_from_csv(path)

    specs: List[DatasetSpec] = []
    header = None
    seen_names = {}
    for raw in rows:
        cells = ["" if c is None else str(c).strip() for c in raw]
        if not any(cells):
            continue
        if cells[0].startswith("#"):
            continue
        if header is None:
            header = _norm_header(cells)
            if "base_dir" not in header:
                raise ValueError(
                    "Manifest %r needs a base-directory column (one of: %s). "
                    "Got header: %s" % (path, ", ".join(_ALIASES["base_dir"]), cells)
                )
            continue

        def cell(canon):
            i = header.get(canon)
            return cells[i] if i is not None and i < len(cells) else ""

        base = cell("base_dir")
        if not base:
            continue
        if not os.path.isabs(base):
            base = os.path.normpath(os.path.join(manifest_dir, base))
        cfg = cell("config") or None
        if cfg and not os.path.isabs(cfg):
            cfg = os.path.normpath(os.path.join(manifest_dir, cfg))
        name = cell("name") or os.path.basename(base.rstrip("/\\")) or base
        # disambiguate duplicate names so state keys and logs stay unique
        if name in seen_names:
            seen_names[name] += 1
            name = "%s#%d" % (name, seen_names[name])
        else:
            seen_names[name] = 0
        specs.append(DatasetSpec(name=name, base_dir=base, config=cfg))
    if header is None:
        raise ValueError("Manifest %r has no header row" % path)
    return specs


# --------------------------------------------------------------------------- #
# Signature (for resume) + pre-flight
# --------------------------------------------------------------------------- #
def _tif_leaf_dirs(base_dir):
    """Yield leaf directories that look like a movie (contain .tif or filXYs)."""
    for root, subdirs, files in os.walk(base_dir):
        if subdirs:
            continue
        if any(f.endswith(".tif") for f in files) or any(
            f.startswith("filXYs") for f in files
        ):
            yield root, files


def _input_fingerprint(base_dir):
    """Cheap (count, total_bytes, latest_mtime) over the dataset's tif files."""
    count = 0
    total = 0
    latest = 0.0
    for root, _subdirs, files in os.walk(base_dir):
        for f in files:
            if not f.endswith(".tif"):
                continue
            try:
                st = os.stat(os.path.join(root, f))
            except OSError:
                continue
            count += 1
            total += st.st_size
            latest = max(latest, st.st_mtime)
    return count, total, latest


def signature(spec: DatasetSpec) -> str:
    """A short signature of the inputs + config; changes when either changes.

    Uses a lightweight fingerprint (file count / total size / newest mtime) rather
    than hashing gigabytes of image data, plus the config file's bytes.
    """
    h = hashlib.sha1()
    h.update(os.path.abspath(spec.base_dir).encode())
    h.update(repr(_input_fingerprint(spec.base_dir)).encode())
    if spec.config and os.path.isfile(spec.config):
        with open(spec.config, "rb") as f:
            h.update(f.read())
    return h.hexdigest()[:16]


def _run_kwargs_for(spec: DatasetSpec) -> Dict:
    """Build gliding.run kwargs from the dataset's config (or defaults)."""
    from ..config import Settings
    settings = Settings.from_toml(spec.config) if spec.config else Settings()
    return settings.to_run_kwargs()


def preflight(spec: DatasetSpec, smoke: bool = False) -> List[str]:
    """Return a list of problems for ``spec`` (empty == ready to run).

    Structural checks by default; ``smoke=True`` additionally detects the first
    frame of the first movie to catch detector/dependency errors early.
    """
    problems: List[str] = []
    if not os.path.isdir(spec.base_dir):
        problems.append("base directory not found: %s" % spec.base_dir)
        return problems  # nothing else is checkable

    leaves = list(_tif_leaf_dirs(spec.base_dir))
    if not leaves:
        problems.append("no movie folders with .tif (or cached filXYs) found under %s"
                        % spec.base_dir)

    run_kwargs = None
    if spec.config and not os.path.isfile(spec.config):
        problems.append("config file not found: %s" % spec.config)
    else:
        try:
            run_kwargs = _run_kwargs_for(spec)
        except Exception as exc:
            problems.append("config failed to parse (%s): %s" % (spec.config, exc))

    # detector optional dependency present?
    if run_kwargs is not None:
        algo = run_kwargs.get("detection_algorithm", "entropy")
        mod = {"ridge": "ridge_detector", "ridge-fast": "ridge_detector_fast"}.get(algo)
        if mod is not None:
            import importlib.util
            if importlib.util.find_spec(mod) is None:
                problems.append(
                    "detector %r needs the optional package %r (pip install "
                    "'fastrack[%s]')" % (algo, mod, algo))

    # output dir writable (we write under ./outputs)
    try:
        os.makedirs("outputs", exist_ok=True)
        if not os.access("outputs", os.W_OK):
            problems.append("output directory ./outputs is not writable")
    except OSError as exc:
        problems.append("cannot create ./outputs: %s" % exc)

    if smoke and not problems and leaves:
        problems.extend(_smoke_detect(spec, leaves[0], run_kwargs or {}))
    return problems


def _smoke_detect(spec, leaf, run_kwargs):
    """Detect frame 0 of one movie; report any error (slow; opt-in)."""
    root, files = leaf
    tifs = sorted(f for f in files if f.endswith(".tif"))
    if not tifs:
        return []
    try:
        from ..core.frame import Frame
        from ..core.detection import DETECTORS
        tail = tifs[0].split("_")[2] if len(tifs[0].split("_")) > 2 else ""
        fr = Frame()
        fr.directory, fr.header, fr.tail = root, "img_000000", tail
        if not fr.read_frame(0):
            return ["smoke: could not read frame 0 in %s" % root]
        algo = run_kwargs.get("detection_algorithm", "entropy")
        params = run_kwargs.get("detection_params", {}) or {}
        if algo == "entropy":
            det = DETECTORS.create("entropy",
                                   fast_rank=run_kwargs.get("fast_rank", True),
                                   morph_contrast=run_kwargs.get("morph_contrast", False))
        else:
            det = DETECTORS.create(algo, **params)
        det.detect(fr)
    except Exception as exc:
        return ["smoke: detection failed on %s: %s" % (root, exc)]
    return []


# --------------------------------------------------------------------------- #
# State
# --------------------------------------------------------------------------- #
def load_state(path):
    if path and os.path.isfile(path):
        try:
            with open(path) as f:
                return json.load(f)
        except (OSError, ValueError):
            LOGGER.warning("could not read state file %s; starting fresh", path)
    return {"datasets": {}}


def save_state(path, state):
    tmp = path + ".tmp"
    with open(tmp, "w") as f:
        json.dump(state, f, indent=2, sort_keys=True)
    os.replace(tmp, path)


# --------------------------------------------------------------------------- #
# Logging
# --------------------------------------------------------------------------- #
def _setup_logging(logdir, verbose):
    os.makedirs(logdir, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    master = os.path.join(logdir, "batch_%s.log" % stamp)
    LOGGER.handlers.clear()
    LOGGER.setLevel(logging.DEBUG)
    fmt = logging.Formatter("%(asctime)s %(levelname)-7s %(message)s",
                            "%Y-%m-%d %H:%M:%S")
    fh = logging.FileHandler(master)
    fh.setLevel(logging.DEBUG)
    fh.setFormatter(fmt)
    LOGGER.addHandler(fh)
    ch = logging.StreamHandler()
    ch.setLevel(logging.DEBUG if verbose else logging.INFO)
    ch.setFormatter(fmt)
    LOGGER.addHandler(ch)
    LOGGER.propagate = False
    return master


# --------------------------------------------------------------------------- #
# Driver
# --------------------------------------------------------------------------- #
def run_batch(manifest, state=None, logdir="fastrack_batch_logs", force=False,
              retry_failed=False, preflight_only=False, smoke=False,
              nprocs=None, stop_on_error=False, verbose=False):
    """Process every dataset in ``manifest``; never stop on a single failure."""
    master_log = _setup_logging(logdir, verbose)
    if state is None:
        state = os.path.join(logdir, "batch_state.json")

    LOGGER.info("FASTrack batch run | manifest=%s | log=%s", manifest, master_log)
    specs = read_manifest(manifest)
    LOGGER.info("manifest lists %d dataset(s)", len(specs))

    # ----- pre-flight over the whole list ------------------------------------ #
    LOGGER.info("running pre-flight checks%s ...", " (with smoke detect)" if smoke else "")
    n_ok = 0
    for spec in specs:
        spec.problems = preflight(spec, smoke=smoke)
        if spec.problems:
            LOGGER.warning("PREFLIGHT FAIL  %-30s", spec.name)
            for p in spec.problems:
                LOGGER.warning("    - %s", p)
        else:
            n_ok += 1
            LOGGER.debug("preflight ok    %s", spec.name)
    LOGGER.info("pre-flight: %d/%d dataset(s) ready", n_ok, len(specs))
    if preflight_only:
        LOGGER.info("pre-flight only; exiting without processing")
        return _summary(specs, {}, preflight_only=True)

    # ----- process ----------------------------------------------------------- #
    st = load_state(state)
    datasets = st.setdefault("datasets", {})
    results = {}
    t0 = time.time()
    for i, spec in enumerate(specs, 1):
        prev = datasets.get(spec.name, {})
        LOGGER.info("[%d/%d] %s", i, len(specs), spec.name)

        if spec.problems:
            LOGGER.warning("  skipping (pre-flight failed)")
            datasets[spec.name] = {"status": "preflight_failed", "problems": spec.problems,
                                   "finished": _now()}
            results[spec.name] = "preflight_failed"
            save_state(state, st)
            continue

        sig = signature(spec)
        if not force and prev.get("status") == "done" and prev.get("signature") == sig:
            LOGGER.info("  already done (unchanged) -> skip")
            results[spec.name] = "skipped"
            continue
        if (not force and not retry_failed and prev.get("status") == "failed"
                and prev.get("signature") == sig):
            LOGGER.info("  previously failed (unchanged); use --retry-failed to retry -> skip")
            results[spec.name] = "skipped_failed"
            continue

        ds_log = os.path.join(logdir, "%s.log" % _safe(spec.name))
        LOGGER.info("  running -> %s  (detail log: %s)", spec.base_dir, ds_log)
        status, error = _run_one(spec, nprocs, ds_log)
        entry = {"status": status, "signature": sig, "base_dir": os.path.abspath(spec.base_dir),
                 "config": spec.config, "finished": _now(), "detail_log": ds_log}
        if error:
            entry["error"] = error
        datasets[spec.name] = entry
        results[spec.name] = status
        save_state(state, st)  # crash-safe: persist after every dataset

        if status == "done":
            LOGGER.info("  OK")
        else:
            LOGGER.error("  FAILED: %s", error)
            if stop_on_error:
                LOGGER.error("stop_on_error set; aborting remaining datasets")
                break

    LOGGER.info("batch finished in %.1f s", time.time() - t0)
    return _summary(specs, results)


def _run_one(spec, nprocs, ds_log):
    """Run one dataset; capture its (chatty) output to ``ds_log``. Never raises."""
    try:
        run_kwargs = _run_kwargs_for(spec)
    except Exception as exc:
        return "failed", "config error: %s" % exc
    if nprocs is not None:
        run_kwargs["nprocs"] = nprocs

    t0 = time.time()
    try:
        from . import gliding  # lazy: pulls in the image stack
        with open(ds_log, "w") as logf:
            with redirect_stdout(logf), redirect_stderr(logf):
                gliding.run(main_dir=spec.base_dir, **run_kwargs)
    except KeyboardInterrupt:
        raise
    except (Exception, SystemExit) as exc:
        # SystemExit too: gliding.run calls sys.exit() on a bad directory, which
        # must not kill the whole batch.
        tb = traceback.format_exc()
        try:
            with open(ds_log, "a") as logf:
                logf.write("\n=== FAILED after %.1f s ===\n%s\n" % (time.time() - t0, tb))
        except OSError:
            pass
        return "failed", "%s: %s" % (type(exc).__name__, exc)
    return "done", None


def _summary(specs, results, preflight_only=False):
    from collections import Counter
    counts = Counter(results.values())
    if preflight_only:
        ok = sum(1 for s in specs if not s.problems)
        LOGGER.info("SUMMARY (pre-flight): %d ready, %d with problems",
                    ok, len(specs) - ok)
    else:
        LOGGER.info("SUMMARY: %s", dict(counts))
    return {"results": dict(results),
            "counts": dict(counts),
            "preflight_problems": {s.name: s.problems for s in specs if s.problems}}


def _now():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def _safe(name):
    return "".join(c if c.isalnum() or c in "-_." else "_" for c in name)
